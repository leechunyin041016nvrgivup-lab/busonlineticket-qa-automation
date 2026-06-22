"""
utils/data_loader.py
────────────────────
Loads test data from data/test_data.xlsx (primary source).

Sheet layout — "TestData":
  Row 1 : headers  →  Section | Key | Value
  Row 2+: data rows (one setting per row)

If the Excel file does not exist yet it is auto-created from
data/test_data.json so existing projects migrate seamlessly.
"""

from __future__ import annotations
import json
import re
from pathlib import Path
from config.settings import DATA_FILE, EXCEL_FILE


# Column on the BookingTestData sheet that holds the full per-case booking
# payload as a single JSON object (passengers, meals, payment, expected fields).
BOOKING_JSON_COL = "Test Data (JSON)"


# Fallback passenger-detail fill values, used when the TestData sheet of an
# existing workbook predates the "manifest" section.
_DEFAULT_MANIFEST = {
    "full_name":       "QA Tester",
    "phone":           "0163553613",
    "email":           "qa.tester@example.com",
    "gender":          "Male",
    "dob":             "01/01/1995",
    "nationality":     "Malaysia",
    "passport_no":     "A12345678",
    "passport_expiry": "01/01/2032",
}

# Per-case booking payloads seeded into the "Test Data (JSON)" cell of each
# BookingTestData row. Lead Full Name/Phone/Email are always the fixed dummy
# (LEECHUNYIN/0163553613/leechunyin@gmail.com— see the test module), so
# they are NOT repeated here. `expected_fields` is left empty (record-only): the
# report records which manifest fields each route actually shows, without a
# strict pass/fail that could misfire before the flow is proven end-to-end.
_JSON_TRANSTAR_BUS = json.dumps({
    "passengers": [
        {"first_name": "Test", "last_name": "Batch Five", "gender": "Male",
         "dob": "1998-01-01", "nationality": "Germany",
         "passport_no": "A12348866", "passport_expiry": "2028-04-01"},
        {"first_name": "Test", "last_name": "Second", "gender": "Male",
         "dob": "2006-01-01", "nationality": "Malaysia",
         "passport_no": "A12345555", "passport_expiry": "2029-01-01"},
    ],
    "depart_meal": {"vegetarian": 2, "non_vegetarian": 0},
    "return_meal": {"vegetarian": 1, "non_vegetarian": 1},
    "addons": ["insurance", "refund"],
    "discount": "12.50", "total": "1148.89", "currency": "RM",
    "payment_method": "Touch N Go",
    "expected_fields": [],
}, ensure_ascii=False, indent=2)

_JSON_BINTAN_FERRY = json.dumps({
    "passengers": [
        {"first_name": "Test", "last_name": "Batch Five", "gender": "Male",
         "dob": "2004-01-01", "nationality": "Cambodia",
         "passport_no": "A12346677", "passport_expiry": "2029-01-01",
         "country_of_issuance": "Cambodia"},
        {"first_name": "Test", "last_name": "Second", "gender": "Male",
         "dob": "2026-01-01", "nationality": "Singapore",
         "passport_no": "A12348324", "passport_expiry": "2036-01-01",
         "country_of_issuance": "Singapore"},
    ],
    "child": 1,
    "addons": ["promo"],
    "discount": "66.98", "total": "646.31", "currency": "RM",
    "payment_method": "Boost",
    "expected_fields": [],
}, ensure_ascii=False, indent=2)

_JSON_DOLPHIN_FERRY = json.dumps({
    "passengers": [
        {"first_name": "Test", "last_name": "Batch Five", "gender": "Male",
         "passenger_type": "Adult", "dob": "2003-04-01", "nationality": "Greece",
         "passport_no": "A12349976", "passport_issue": "2020-04-01",
         "passport_expiry": "2028-04-01", "country_of_issuance": "Greece"},
        {"first_name": "Test", "last_name": "Second", "gender": "Male",
         "passenger_type": "Child", "dob": "2025-01-01", "nationality": "Singapore",
         "passport_no": "A12346636", "passport_issue": "2026-04-01",
         "passport_expiry": "2026-07-01", "country_of_issuance": "Singapore"},
    ],
    "child": 1,
    "addons": ["refund", "promo"],
    "discount": "2.00", "total": "110.70", "currency": "SGD",
    "payment_method": "PayNow",
    "expected_fields": [],
}, ensure_ascii=False, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
#  Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _load_from_json() -> dict:
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _create_excel_from_json(data: dict) -> None:
    """Bootstrap the Excel file from the JSON seed data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── TestData sheet ────────────────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "TestData"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    for col, heading in enumerate(["Section", "Key", "Value"], start=1):
        cell = ws_data.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws_data.column_dimensions["A"].width = 18
    ws_data.column_dimensions["B"].width = 22
    ws_data.column_dimensions["C"].width = 45

    row = 2
    for section, fields in data.items():
        if isinstance(fields, dict):
            for key, value in fields.items():
                ws_data.cell(row=row, column=1, value=section)
                ws_data.cell(row=row, column=2, value=key)
                ws_data.cell(row=row, column=3, value=str(value) if value is not None else "")
                row += 1
        else:
            ws_data.cell(row=row, column=1, value=section)
            ws_data.cell(row=row, column=2, value="value")
            ws_data.cell(row=row, column=3, value=str(fields))
            row += 1

    # ── TestResults sheet ─────────────────────────────────────────────────────
    ws_results = wb.create_sheet("TestResults")

    result_headers = ["Run Date", "TC ID", "Test Name", "Type", "Status", "Duration (s)", "Error Message"]
    for col, heading in enumerate(result_headers, start=1):
        cell = ws_results.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws_results.column_dimensions["A"].width = 20
    ws_results.column_dimensions["B"].width = 14
    ws_results.column_dimensions["C"].width = 40
    ws_results.column_dimensions["D"].width = 8
    ws_results.column_dimensions["E"].width = 10
    ws_results.column_dimensions["F"].width = 14
    ws_results.column_dimensions["G"].width = 60

    _seed_booking_data(wb)

    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)


def _seed_booking_data(wb) -> None:
    """Add the BookingTestData sheet with initial test cases."""
    from openpyxl.styles import Font, PatternFill, Alignment

    if "BookingTestData" in wb.sheetnames:
        return  # already exists, don't overwrite

    ws = wb.create_sheet("BookingTestData")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    # "Test Data (JSON)" holds the full per-case booking payload as one JSON
    # object: passengers[], depart_meal/return_meal, lead phone/email, discount,
    # total, payment_method, and an optional expected_fields[] list for the
    # manifest visibility check. See parse_booking_json() for the schema.
    headers = ["TC ID", "Transport", "Operator", "Origin", "Destination",
               "Depart Date", "Return Date", BOOKING_JSON_COL, "Status"]
    col_widths = [12, 10, 28, 22, 22, 14, 14, 60, 10]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    seed_rows = [
        ("TC-BK-01", "Bus",   "Transtar Travel Express",       "Singapore", "Genting Highlands", "30/06/2026", "", _JSON_TRANSTAR_BUS,  "active"),
        ("TC-BK-03", "Ferry", "Bintan Resort Ferry (Economy)", "Singapore", "Bintan",            "30/06/2026", "", _JSON_BINTAN_FERRY,  "active"),
        ("TC-BK-04", "Ferry", "Dolphin Fast Ferry",            "Stulang",   "Batam",             "30/06/2026", "", _JSON_DOLPHIN_FERRY, "active"),
    ]
    for r, row_data in enumerate(seed_rows, start=2):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)


def _ensure_booking_columns() -> None:
    """
    Idempotently make sure the BookingTestData sheet has the "Test Data (JSON)"
    column. Renames a pre-existing "Expected Fields" column to it (so workbooks
    from the earlier comma-list design migrate in place), or adds it before the
    Status column. Safe to call on every read; swallows the PermissionError
    raised when the file is open in Excel.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        wb = load_workbook(EXCEL_FILE)
        if "BookingTestData" not in wb.sheetnames:
            return
        ws = wb["BookingTestData"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if BOOKING_JSON_COL in headers:
            return

        # Migrate the old comma-list column in place if present.
        if "Expected Fields" in headers:
            col = headers.index("Expected Fields") + 1
            ws.cell(row=1, column=col, value=BOOKING_JSON_COL)
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 60
            wb.save(EXCEL_FILE)
            return

        # Otherwise insert before Status (or append).
        if "Status" in headers:
            target = headers.index("Status") + 1   # 1-based index of Status
            ws.insert_cols(target)
        else:
            target = len(headers) + 1

        cell = ws.cell(row=1, column=target, value=BOOKING_JSON_COL)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 60
        wb.save(EXCEL_FILE)
    except PermissionError:
        # File is open in Excel — skip migration this run.
        pass
    except Exception:
        pass


def _load_from_excel() -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["TestData"]

    data: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        section, key, value = row[0], row[1], row[2]
        if section and key:
            data.setdefault(str(section), {})[str(key)] = value if value is not None else ""
    return data


# ─────────────────────────────────────────────────────────────────────────────
#  Public API
# ─────────────────────────────────────────────────────────────────────────────

def load_test_data() -> dict:
    """Return the full test data dict, sourcing from Excel (creates it if missing)."""
    if not EXCEL_FILE.exists():
        seed = _load_from_json() if DATA_FILE.exists() else {}
        _create_excel_from_json(seed)

    return _load_from_excel()


def get_login_data() -> dict:
    return load_test_data()["login"]


def get_signup_data() -> dict:
    return load_test_data()["signup"]


def get_base_urls() -> dict:
    return load_test_data()["base_url"]


def _norm_date(value) -> str:
    """
    Normalise a date to DD/MM/YYYY (the format the site's datepicker expects).
    Accepts YYYY-MM-DD / YYYY/MM/DD (ISO) and DD/MM/YYYY / DD-MM-YYYY.
    Leaves anything unrecognised untouched.
    """
    s = str(value or "").strip()
    if not s:
        return ""
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", s)        # ISO: yyyy-mm-dd
    if m:
        y, mo, d = m.groups()
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", s)        # dd/mm/yyyy
    if m:
        d, mo, y = m.groups()
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    return s


def parse_booking_json(raw: str) -> dict:
    """
    Parse the "Test Data (JSON)" cell into a normalised booking payload.

    Expected JSON shape (all keys optional):
        {
          "passengers": [
            {"first_name","last_name","gender","dob","nationality",
             "passport_no","passport_expiry",
             # extra ferry/train fields — stored, filled only if a catalog
             # field exists for them:
             "passenger_type","passport_issue","country_of_issuance"}, ...
          ],
          "child": 1,                       # No. of child tickets (lead block)
          "depart_meal": {"vegetarian": int, "non_vegetarian": int},
          "return_meal": {"vegetarian": int, "non_vegetarian": int},
          "addons": ["insurance","refund","promo","qr"],   # add-ons to tick
          "lead":        {"phone","email"},
          "discount": "12.50", "total": "1148.89", "currency": "RM",
          "payment_method": "Touch N Go",
          "expected_fields": ["lead_name", "other_name", ...]
        }

    Dates are normalised to DD/MM/YYYY and each passenger gets a derived
    "full_name" = "First Last". On bad JSON, result["error"] is set (no raise).
    """
    result = {
        "passengers": [], "child": "", "depart_meal": {}, "return_meal": {},
        "addons": [], "lead": {},
        "discount": "", "total": "", "currency": "", "payment_method": "",
        "expected_fields": [], "error": None,
    }
    raw = str(raw or "").strip()
    if not raw:
        return result

    try:
        data = json.loads(raw)
    except Exception as e:
        result["error"] = f"Invalid JSON: {e}"
        return result
    if not isinstance(data, dict):
        result["error"] = "Test Data JSON must be a JSON object {...}"
        return result

    for p in (data.get("passengers") or []):
        if not isinstance(p, dict):
            continue
        fn = str(p.get("first_name", "")).strip()
        ln = str(p.get("last_name", "")).strip()
        full = (f"{fn} {ln}").strip() or str(p.get("full_name", "")).strip()
        result["passengers"].append({
            "first_name": fn, "last_name": ln, "full_name": full,
            "gender":          str(p.get("gender", "")).strip(),
            "dob":             _norm_date(p.get("dob", "")),
            "nationality":     str(p.get("nationality", "")).strip(),
            "passport_no":     str(p.get("passport_no", "")).strip(),
            "passport_expiry": _norm_date(p.get("passport_expiry", "")),
            # extra fields (filled only where a manifest field exists)
            "passenger_type":      str(p.get("passenger_type", "")).strip(),
            "passport_issue":      _norm_date(p.get("passport_issue", "")),
            "country_of_issuance": str(p.get("country_of_issuance", "")).strip(),
        })

    result["child"]         = str(data.get("child", "")).strip()
    result["depart_meal"]   = data.get("depart_meal") or {}
    result["return_meal"]   = data.get("return_meal") or {}
    result["addons"]        = [_norm_addon(x) for x in (data.get("addons") or []) if _norm_addon(x)]
    result["lead"]          = data.get("lead") or {}
    result["discount"]      = str(data.get("discount", "")).strip()
    result["total"]         = str(data.get("total", "")).strip()
    result["currency"]      = str(data.get("currency", "")).strip()
    result["payment_method"] = str(data.get("payment_method", "")).strip()
    result["expected_fields"] = [
        str(x).strip() for x in (data.get("expected_fields") or []) if str(x).strip()
    ]
    return result


# Canonical add-on keys understood by the manifest filler. Maps the many ways a
# tester might name an add-on to one of: insurance / refund / promo / qr.
def _norm_addon(value) -> str:
    s = str(value or "").strip().lower()
    if not s:
        return ""
    if "insur" in s:
        return "insurance"
    if "refund" in s or "protect" in s:
        return "refund"
    if "promo" in s or "featured" in s or "pop" in s:
        return "promo"
    if "qr" in s or "boarding" in s:
        return "qr"
    return ""


def get_manifest_data() -> dict:
    """
    Return the passenger-detail values used to auto-fill the booking manifest
    (payment) page. Reads the 'manifest' section of the TestData sheet, falling
    back to _DEFAULT_MANIFEST for any key the sheet doesn't provide.
    """
    data = load_test_data()
    manifest = data.get("manifest", {}) or {}
    merged = dict(_DEFAULT_MANIFEST)
    merged.update({str(k): str(v) for k, v in manifest.items() if v not in (None, "")})
    return merged


def get_booking_data() -> list[dict]:
    """Return booking test cases from the BookingTestData sheet (active rows only)."""
    if not EXCEL_FILE.exists():
        load_test_data()  # triggers file + sheet creation

    from openpyxl import load_workbook

    wb = load_workbook(EXCEL_FILE, data_only=True)

    # Add the sheet on-the-fly if the file predates this feature
    if "BookingTestData" not in wb.sheetnames:
        wb2 = load_workbook(EXCEL_FILE)
        _seed_booking_data(wb2)
        wb2.save(EXCEL_FILE)
        wb = load_workbook(EXCEL_FILE, data_only=True)

    # Add the "Expected Fields" column for workbooks that predate it.
    _ensure_booking_columns()
    wb = load_workbook(EXCEL_FILE, data_only=True)

    ws = wb["BookingTestData"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        d = {str(h): (str(v) if v is not None else "") for h, v in zip(headers, row)}
        # Treat a blank Status as active — only skip rows EXPLICITLY marked
        # inactive (e.g. "skip"/"inactive"/"disabled"). The "Status" header is
        # always present, so a default in .get() would never apply to a blank cell.
        status = d.get("Status", "").strip().lower()
        if status in ("", "active"):
            rows.append(d)
    return rows
