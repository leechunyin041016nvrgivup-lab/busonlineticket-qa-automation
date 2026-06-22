"""
utils/excel_writer.py
─────────────────────
Appends test run results to the "TestResults" sheet in data/test_data.xlsx.

Each test run adds rows at the bottom so history is preserved across runs.
Columns: Run Date | TC ID | Test Name | Type | Status | Duration (s) | Error Message
"""

from __future__ import annotations
import datetime
import re
from pathlib import Path

from config.settings import EXCEL_FILE


def _case_id(result: dict) -> str:
    """
    Best per-case id for a booking test. Parametrized tests share one docstring
    so result['tc_id'] is the generic 'TC-BK-XX'; the real id (TC-BK-01) lives in
    the bracketed parametrize id of result['name'].
    """
    m = re.search(r"\[(.*?)\]", result.get("name", ""))
    return m.group(1) if m else result.get("tc_id", "")


def append_results(ui_results: list[dict], api_results: list[dict]) -> None:
    """Append UI and API test results to the TestResults sheet."""
    if not EXCEL_FILE.exists():
        return  # data_loader creates the file; skip if somehow missing

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return

    wb = load_workbook(EXCEL_FILE)

    if "TestResults" not in wb.sheetnames:
        return

    ws = wb["TestResults"]

    run_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    pass_font  = Font(color="166534")
    fail_font  = Font(color="991B1B")
    pass_fill  = PatternFill("solid", fgColor="DCFCE7")
    fail_fill  = PatternFill("solid", fgColor="FEE2E2")

    def _write_row(result: dict, test_type: str):
        row = ws.max_row + 1
        status = "PASS" if result["passed"] else "FAIL"
        error  = (result.get("error_message") or "").replace("\n", " ")[:500]

        ws.cell(row=row, column=1, value=run_date)
        ws.cell(row=row, column=2, value=result.get("tc_id", ""))
        ws.cell(row=row, column=3, value=result.get("name", ""))
        ws.cell(row=row, column=4, value=test_type)
        status_cell = ws.cell(row=row, column=5, value=status)
        ws.cell(row=row, column=6, value=round(result.get("duration", 0), 2))
        ws.cell(row=row, column=7, value=error)

        if result["passed"]:
            status_cell.font = pass_font
            status_cell.fill = pass_fill
        else:
            status_cell.font = fail_font
            status_cell.fill = fail_fill

    for r in ui_results:
        _write_row(r, "UI")
    for r in api_results:
        _write_row(r, "API")

    wb.save(EXCEL_FILE)


def append_manifest_results(ui_results: list[dict]) -> int:
    """
    Append per-field manifest visibility results to the "ManifestResults" sheet.

    One row per (test case × manifest field):
      Run Date | TC ID | Test Name | Field Key | Field | Shown | Expected | Verdict

    Verdict colouring: PASS green, FAIL red, '—' (record-only) left plain.
    Returns the number of rows written.
    """
    if not EXCEL_FILE.exists():
        return 0

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return 0

    wb = load_workbook(EXCEL_FILE)

    headers = ["Run Date", "TC ID", "Test Name", "Field Key", "Field", "Shown", "Expected", "Verdict"]
    col_widths = [20, 12, 42, 22, 30, 8, 10, 9]

    if "ManifestResults" not in wb.sheetnames:
        ws = wb.create_sheet("ManifestResults")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = w
    else:
        ws = wb["ManifestResults"]

    run_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    pass_font = Font(color="166534")
    fail_font = Font(color="991B1B")
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")

    written = 0
    for r in ui_results:
        fields = r.get("manifest_fields") or []
        if not fields:
            continue
        case_id   = _case_id(r)
        test_name = r.get("name", "")
        for f in fields:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=run_date)
            ws.cell(row=row, column=2, value=case_id)
            ws.cell(row=row, column=3, value=test_name)
            ws.cell(row=row, column=4, value=f.get("key", ""))
            ws.cell(row=row, column=5, value=f.get("label", ""))
            ws.cell(row=row, column=6, value="Yes" if f.get("shown") else "No")
            ws.cell(row=row, column=7, value=f.get("expected", "—"))
            verdict_cell = ws.cell(row=row, column=8, value=f.get("verdict", "—"))

            if f.get("verdict") == "PASS":
                verdict_cell.font = pass_font
                verdict_cell.fill = pass_fill
            elif f.get("verdict") == "FAIL":
                verdict_cell.font = fail_font
                verdict_cell.fill = fail_fill
            written += 1

    if written:
        wb.save(EXCEL_FILE)
    return written
